# some tests. This page is not to be used on its own

# For working with Excel, install the OpenPyXl module, pip install --user openpyxl.
# More info at https://openpyxl.readthedocs.io/en/stable/


# 0. Imports
from fileinput import filename
import openpyxl
from openpyxl.utils import get_column_letter

# MULTIPLICATION TABLE MAKER, p. 326, 'AUtomate the boring stuff with python', Al Seigwart
print('opening workbook')
fileName = 'multiplication-table.xlsx'
try:
    wb = openpyxl.Workbook()
    sheet = wb.active

    print('calculating...')
    number = 6 # TODO: get this from command line
    for i in range(1, number + 1):
        #  First row, print number 1...x
        sheet.cell(row=1, column=i+1).value= i
        # second row, print number again, the i * numbers
        newRow = [i,]
        # calculate the multiplication number
        for j in range(1, number+ 1):
            newRow.append(j * i)
        # print to the screen, for reference, and add to worksheet
        print('adding row: ', newRow)
        sheet.append(newRow)

except Exception as ex:
    print('ERROR! {0}'.format(ex))

print('closing workbook: ', fileName)
wb.save(filename=fileName)
print('Done.')