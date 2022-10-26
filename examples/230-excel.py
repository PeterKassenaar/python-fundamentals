# For working with Excel, install the OpenPyXl module, pip install --user openpyxl.
# More info at https://openpyxl.readthedocs.io/en/stable/


# 0. Imports
from fileinput import filename
import openpyxl
from openpyxl.utils import get_column_letter

# 1. We use the Sample.xlsx file with two worksheets and some data.
# This project illustrates some options of the openpyxl module.
print('opening workbook')
fileName = 'hardware_store.xlsx'
try:
    # ************
    # Reading from a workbook
    # ************
    # 2. Opening the workbook
    wb = openpyxl.load_workbook(fileName)

    # # 3. print the sheet names in the workbook:
    # print(wb.sheetnames)

    # # 4. reading the value of a cell. Look for yourself what happens if you
    # # pass in a non-existing sheet or cell.
    # sheet = wb['Info']
    # print(sheet['A1'].value)

    # # 5. You can also store the cell information in a variable
    # cell = sheet['A2']
    # print(type(cell))

    # # 6. getting the row, column and value from the cell.
    # # Note: the column is also an integer, not a Letter.
    # print('Row: {0}, column: {1}, value: {2}'.format(
    #     cell.row, cell.column, cell.value
    # ))

    # # 7. getting the coordinates of the cell
    # print(cell.coordinate)  # A2

    # # 8. Printing all information in the rows
    # sheet_max_row = sheet.max_row
    # sheet_max_column = sheet.max_column
    # # We only traverse the A-column (= first column) here.
    # # Rows start with 1, not with 0!.
    # print('info in first column: ')
    # for i in range(1, sheet_max_row+1):
    #     print('{0}{1}: {2}'.format(
    #         # get the column letter from the 'utils' module
    #         get_column_letter(1),
    #         i,  # index
    #         sheet.cell(row=i, column=1).value))  # get the value

    # #  9. Getting all cells in a rectangular area
    # print('********* rectangular area  **********')
    # for rowOfCells in sheet['A1': 'D4']:
    #     #  Uncomment the next line to see what rowOfCells looks like (its a tuple with the 4 cells).
    #     # print( '( rowOfCells contains a tuple: ', rowOfCells, ')')
    #     for singleCell in rowOfCells:
    #         print(singleCell.coordinate, singleCell.value)
    #     print('********* end row **********')

    # ************
    # Writing to a workbook
    # ************
    #  Simple writing of values:
    # Set the active sheet
    # sheet= wb.active 
    # sheet['A10'] = 'Hello World!'
    #  Get the data back
    #  value = sheet['A10'].value
    # wb.save(fileName)

    # we have already opened the workbook.
    # Now we create a new sheet and write employee names to it:
    # employees = [
    #     ['Peter', 'peter@peter.com'],
    #     ['Sandra', 'sandra@sandra.com'],
    #     ['Bobby', 'bobby@bobby.com']
    # ]
    # # by default, it will be inserted as the last sheet. Use index=... parameter for a specific position.
    # wb.create_sheet('Employees')
    # sheet = wb['Employees']

    # # writing the data
    # for emp in employees:
    #     sheet.append(emp)

    # # Done.
    # print('saving & closing workbook')
    # wb.save(fileName)
    # print('Done.')

    # ************
    #  Case: updating the prices in a worksheet with 15% if product is 'Hammer' or 'Drill'
    # ************
    productsToUpdate = ['Hammer', 'Drill']
    percentageToUpdate = 15
    sheet = wb['Products']

    # loop over all products
    for row in range(1, sheet.max_row + 1):
        # Get the product name
        productName = sheet.cell(row=row, column=1).value
        # IF we need to update this product...
        if productName in productsToUpdate:   
            #... get its current value...         
            currentPrice = sheet.cell(row=row, column=3).value

            # ...apply a little math to calculate the new price
            newPrice = round((currentPrice * (percentageToUpdate / 100) + currentPrice), 2)

            # ...and write the value back to the sheet
            sheet.cell(row=row, column=3).value = newPrice
    wb.save(fileName)
    print('Prices updated!')

    # All kinds of other options, see documentation, for instance https://openpyxl.readthedocs.io/en/stable/filters.html

except Exception as ex:
    print('ERROR! {0}'.format(ex))


# Workshop: Create a new workbook with one sheet, write the table of 3 from 1...1000 in the workbook and save the file.
# Workshop: Get the Products sheet from the example workbook and calculate the average price of a tool. Write the average price in cell C10.
# Workshop: Get the Products sheet and update the inventory of the Saw, Screwdriver and Workbench with a value that is provided at the command line.
# Workshop: Multiplication Table maker - see p. 326 book.